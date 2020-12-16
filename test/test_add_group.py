import pytest
from openpyxl import load_workbook

wb = load_workbook('../groups.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
testdata = list()
for i in range(1, 4):
    testdata.append(sheet.cell(row=i, column=1).value)


@pytest.mark.parametrize("group", testdata)
def test_add_group(app, group):
    old_list = app.groups.get_group_list()
    app.groups.add_new_group(group)
    new_list = app.groups.get_group_list()
    old_list.append(group)
    assert sorted(old_list) == sorted(new_list)
